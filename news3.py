from pygooglenews import GoogleNews
import pandas as pd
from openpyxl import load_workbook
import nltk

gn = GoogleNews(lang = 'ar', country = 'sa')
articles=[]
newsitem=[]

search = 'بنك'
newsitem = gn.search(search,when='6m')
newsitem = newsitem['entries']

for item in newsitem:
        article = {
            'title':item.title,
            'link': item.link,
            'Puplished': item.published
                      
        }
        articles.append(article)

df = pd.DataFrame(articles)
print(df)
#df.to_excel('news.xlsx',index=False)

book = load_workbook('news.xlsx')
writer = pd.ExcelWriter('news.xlsx', engine='openpyxl') 
writer.book = book

df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

writer.save()
writer.close() 


nltk.tokenize.wordpunct_tokenize
